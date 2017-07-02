"""
 Created by waldo on 12/22/16
"""

import sys
import sqlite3
import calendar
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from dateutil import rrule
from dateutil.relativedelta import relativedelta
from datetime import date, datetime

COLUMN_ADD_FACTOR = 10
COLUMN_MUL_FACTOR = 1.3
balance_file = '/media/waldo/DATA-SHARE/Code/BudgetCalc/test/balance.xlsx'
prognosis_file = '/media/waldo/DATA-SHARE/Code/BudgetCalc/test/prognosis.xlsx'


class BudgetCalc(object):
    def __init__(self):
        self.calcbook = TransactionWorkbook()

    def read_input(self, filename):
        self.calcbook.load(filename)
        pass

    def save_prognosis(self, folder, years=1, months=0):
        filename = self._compose_filename(folder, years, months)
        self.calcbook.export(filename, years, months)

    @staticmethod
    def _compose_filename(folder, years, months):
        now = date.today()
        then = now + relativedelta(years=+years, months=+months)

        now_str = "{0}{1}".format(calendar.month_name[now.month][:3], now.year)
        then_str = "{0}{1}".format(calendar.month_name[then.month][:3], then.year)

        filename = "{0}/{1}_Budget_{2}-{3}.xlsx".format(folder, str(now).translate(None, '-'), now_str, then_str)
        return filename


class BudgetWorkbook(object):
    def __init__(self):
        self.workbook = Workbook()
        self.current_sheet = self.workbook.active
        self.num_rows = 0
        self.num_cols = 0

    def load(self, filename):
        self.workbook = load_workbook(filename=filename, data_only=True)
        self.workbook.guess_types = True
        self.current_sheet = self.workbook.active
        self.num_rows = self.current_sheet.max_row
        self.num_cols = self.current_sheet.max_column


class TransactionWorkbook(BudgetWorkbook):
    # Input columns
    BANKS_COLS = ["Bank", "Currency", "Current balance", "Date (dd-mm-yyyy)"]
    BALANCES_COLS = ["Bank", "Description", "Subsection", "Currency", "Amount (+/-)", "Date (dd-mm-yyyy)",
                     "Interval (# months)", "Repetitions (#)", "Cuotas (#/#)"]

    # Output columns
    ACCOUNTS_COLS = ["Bank", "Balance at end of month"]
    TRANSACTIONS_COLS = ["Description", "Subsection", "Amount", "Date", "Balance after transaction"]

    def __init__(self):
        super(TransactionWorkbook, self).__init__()
        self.db_connection = None
        self.db_cursor = None
        self.accounts_section = []
        self.transaction_section = []
        self.header_font = Font(name="Calibri", size=11, bold=True)
        self.text_font = Font(name="Calibri", size=11)
        self.grey_font = Font(name="Calibri", size=11, color="808080")
        self.month_to_num = {name: num for num, name in enumerate(calendar.month_abbr)}

    @property
    def accounts_current(self):
        start_row = self._find_row(self.BANKS_COLS) + 1

        balances = dict()
        for index, row in enumerate(self.current_sheet.iter_rows(min_row=start_row,
                                                                 max_col=len(self.BANKS_COLS),
                                                                 max_row=self.num_rows)):
            account_name = row[0].value
            if account_name is not None:
                balance_date = row[3].value

                # If the balance date month is in the past, this account is not processed.
                if relativedelta(balance_date, date.today()).months >= 0:
                    currency = row[1].value
                    balance = row[2].value
                    balances[account_name] = (currency, balance, balance_date, index + 1)
            else:
                break

        return balances

    def load(self, filename):
        super(TransactionWorkbook, self).load(filename)
        self._load_db()

    def export(self, filename, years, months):
        # Sort the accounts according to their order in the source sheet
        sorted_items = sorted(self.accounts_current.items(), key=lambda ac: ac[1][3])
        sorted_accounts = [item[0] for item in sorted_items]

        # Prepare a framework / the headers in the workbook.
        # TODO: we already have to start the bank accounts before we do the print of frame
        transactions_book = self._print_frame(sorted_accounts, self.ACCOUNTS_COLS, self.TRANSACTIONS_COLS, years,
                                              months)

        for bank_nr, account in enumerate(sorted_accounts):
            account_date = self.accounts_current[account][2].date()
            account_balance = self.accounts_current[account][1]
            account_currency = self.accounts_current[account][0]

            # Retrieve the entries from database.
            # Sort the entries on the appropriate columns. Unlimited repetitions first.
            value = (account,)
            transactions = self.db_cursor.execute('SELECT * FROM transactions WHERE bank=? '
                                                  'ORDER BY date ASC, reps DESC', value)

            # Initialize a dictionary where the keys are the sheet names, and the values the current cell coordinates.
            cell_pointer = self.transaction_section[0][:]  # Make a shallow copy
            cell_pointer[0] += bank_nr * (len(self.TRANSACTIONS_COLS) + 1)
            cell_pointer[1] += 3
            sheet_pointers = {key: cell_pointer[:] for key in transactions_book.sheetnames}

            current_balance_inserted = False
            transactions_empty = True

            # Sort the transactions by DAY of the month, ascending, irrespective of the actual month.
            sorted_transactions = sorted(transactions, key=lambda tr: tr[5][-11:-9])
            for index, transaction in enumerate(sorted_transactions):
                transactions_empty = False

                # For each entry calculate the months applicable.
                num_months_in_sheet = len(transactions_book.sheetnames)
                start_month = datetime.strptime(transaction[5],
                                                "%Y-%m-%d %H:%M:%S")  # TODO - do we really need the hour etc.
                                                                      # TODO - gives an error if we have for instance a non-existing date like 30/06/2017
                between_start_and_now = 12 * (start_month.year - datetime.now().year) + \
                                        (start_month.month - datetime.now().month)
                num_months_in_range = num_months_in_sheet - between_start_and_now
                applicable_months = self._calc_applicable_months(transaction, num_months_in_range)

                date_time_value = datetime.strptime(transaction[5], "%Y-%m-%d %H:%M:%S")
                transaction_date = date_time_value.date()

                # For each necessary month make an entry.
                for month_nr, month in enumerate(applicable_months):
                    # Initializations
                    day_num = transaction_date.day
                    month_abbr = month[:3]
                    month_num = self.month_to_num[month_abbr]
                    year = int(month[-4:])

                    while True:
                        try:
                            work_date = transaction_date.replace(day=day_num, month=month_num, year=year)
                            break
                        except ValueError:
                            # We end up here when the day does not exist in the month. I.e. 30 Feb.
                            day_num -= 1

                    cell_pointer = sheet_pointers[month]
                    month_sheet = transactions_book.get_sheet_by_name(month)

                    # Write the date. We do this first, because we might need to insert the current balance first.
                    if account_date.month == month_num and account_date.year == year and \
                                    work_date >= account_date and not current_balance_inserted:
                        # TODO: DRY
                        # Description
                        cell = month_sheet.cell(row=cell_pointer[1], column=cell_pointer[0])
                        cell.value = "CURRENT BALANCE"
                        cell.font = self.   header_font
                        cell.number_format = "General"
                        cell.alignment = Alignment(horizontal="justify")

                        # Subsection and Amount should be empty
                        cell = month_sheet.cell(row=cell_pointer[1], column=cell_pointer[0] + 1)
                        cell.value = ""
                        cell.alignment = Alignment(horizontal="justify")
                        cell = month_sheet.cell(row=cell_pointer[1], column=cell_pointer[0] + 2)
                        cell.value = ""
                        cell.alignment = Alignment(horizontal="justify")

                        # Date
                        cell = month_sheet.cell(row=cell_pointer[1], column=cell_pointer[0] + 3)
                        cell.value = account_date
                        cell.font = self.header_font
                        cell.number_format = 'dd-mm-yyyy'
                        cell.alignment = Alignment(horizontal="justify")

                        # Balance
                        cell = month_sheet.cell(row=cell_pointer[1], column=cell_pointer[0] + 4)
                        cell.value = account_balance
                        cell.font = self.header_font
                        cell.number_format = '#,###.00 [${0}];[RED]-#,###.00 [${0}]'.format(account_currency)
                        cell.alignment = Alignment(horizontal="justify")

                        current_balance_inserted = True
                        cell_pointer[1] += 1
                    if work_date >= account_date:
                        fill_colour = PatternFill("none")
                        current_font = self.text_font

                        # Calculate and write the 'balance after transaction'
                        cell = month_sheet.cell(row=cell_pointer[1], column=cell_pointer[0] + 4)
                        cell_prev_row = month_sheet.cell(row=cell_pointer[1] - 1, column=cell_pointer[0] + 4)
                        cell_with_amount = month_sheet.cell(row=cell_pointer[1], column=cell_pointer[0] + 2)
                        cell.value = '=SUM({0},{1})'.format(cell_prev_row.coordinate, cell_with_amount.coordinate)
                        cell.fill = fill_colour
                        cell.number_format = '#,###.00 [$AR$];[RED]-#,###.00 [$AR$]'
                        cell.font = self.grey_font
                        cell.alignment = Alignment(horizontal="justify")
                    else:
                        # Else, the cell should be empty and filled with a grey background colour
                        fill_colour = PatternFill("solid", fgColor="DDDDDD")
                        current_font = self.grey_font
                        cell = month_sheet.cell(row=cell_pointer[1], column=cell_pointer[0] + 4)
                        cell.fill = fill_colour
                        cell.alignment = Alignment(horizontal="justify")

                    # Write the date
                    cell = month_sheet.cell(row=cell_pointer[1], column=cell_pointer[0] + 3)
                    cell.value = work_date
                    cell.number_format = 'dd-mm-yyyy'
                    cell.fill = fill_colour
                    cell.font = current_font
                    cell.alignment = Alignment(horizontal="justify")

                    # Write the description
                    cell = month_sheet.cell(row=cell_pointer[1], column=cell_pointer[0])
                    cell.value = self._compose_description(month_nr, transaction)
                    cell.number_format = "General"
                    cell.fill = fill_colour
                    cell.font = current_font
                    cell.alignment = Alignment(horizontal="justify")

                    # Write the subsection
                    cell = month_sheet.cell(row=cell_pointer[1], column=cell_pointer[0] + 1)
                    cell.value = transaction[2]
                    cell.number_format = "General"
                    cell.fill = fill_colour
                    cell.font = current_font
                    cell.alignment = Alignment(horizontal="justify")

                    # Write the amount
                    cell = month_sheet.cell(row=cell_pointer[1], column=cell_pointer[0] + 2)
                    cell.value = transaction[4]
                    cell.number_format = '#,###.00 [$AR$];[RED]-#,###.00 [$AR$]'
                    cell.fill = fill_colour
                    cell.font = current_font
                    cell.alignment = Alignment(horizontal="justify")

                    cell_pointer[1] += 1
            else:
                # All transactions were entered, but the balance is on a date greater than any transaction.
                if not current_balance_inserted:
                    sheet_title = "{0} {1}".format(calendar.month_name[account_date.month], account_date.year)

                    try:
                        cell_pointer = sheet_pointers[sheet_title]
                        month_sheet = transactions_book.get_sheet_by_name(sheet_title)
                    except Exception as e:
                        print "Account balance date seems to be in the past."
                        break

                    # TODO: DRY
                    # Description
                    cell = month_sheet.cell(row=cell_pointer[1], column=cell_pointer[0])
                    cell.value = "CURRENT BALANCE"
                    cell.font = self.header_font
                    cell.number_format = "General"
                    cell.alignment = Alignment(horizontal="justify")

                    # Subsection and Amount should be empty
                    cell = month_sheet.cell(row=cell_pointer[1], column=cell_pointer[0] + 1)
                    cell.value = ""
                    cell.alignment = Alignment(horizontal="justify")
                    cell = month_sheet.cell(row=cell_pointer[1], column=cell_pointer[0] + 2)
                    cell.value = ""
                    cell.alignment = Alignment(horizontal="justify")

                    # Date
                    cell = month_sheet.cell(row=cell_pointer[1], column=cell_pointer[0] + 3)
                    cell.value = account_date
                    cell.font = self.header_font
                    cell.number_format = 'dd-mm-yyyy'
                    cell.alignment = Alignment(horizontal="justify")

                    # Balance
                    cell = month_sheet.cell(row=cell_pointer[1], column=cell_pointer[0] + 4)
                    cell.value = account_balance
                    cell.font = self.header_font
                    cell.number_format = '#,###.00 [${0}];[RED]-#,###.00 [${0}]'.format(account_currency)
                    cell.alignment = Alignment(horizontal="justify")

                    cell_pointer[1] += 1

            if not transactions_empty:
                self._connect_sheet_formulae(account, transactions_book, sheet_pointers)

        self._autosize_columns(transactions_book, add_factor=COLUMN_ADD_FACTOR, mul_factor=COLUMN_MUL_FACTOR)

        transactions_book.save(filename)

    def _compose_description(self, month_nr, transaction):
        quote_suffix = ""

        # If there are quotes, then incorporate the numbering in the description
        quotes = transaction[8]
        if quotes != "":
            quote_num, total_quotes = [int(x) for x in quotes.split('/')]
            quote_num += month_nr
            quote_suffix = " ({0}/{1})".format(quote_num, total_quotes)

        description = transaction[1] + quote_suffix
        return description

    def _autosize_columns(self, transactions_book, add_factor=0, mul_factor=1.0):
        for sheet in transactions_book.worksheets:
            for column in sheet.columns:
                max_length = 0
                for cell in column:
                    try:  # Necessary to avoid error on empty cells
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass

                adjusted_width = (max_length + mul_factor) + add_factor
                sheet.column_dimensions[cell.column].width = adjusted_width

    def _connect_sheet_formulae(self, account, transactions_book, sheet_pointers):
        # First determine the sheet to start connecting formulae
        balance_date = self.accounts_current[account][2].date()
        first_sheet_title = "{0} {1}".format(calendar.month_name[balance_date.month], balance_date.year)

        start = False
        for sheet in transactions_book.worksheets:
            if sheet.title == first_sheet_title:
                start = True
                pointer = sheet_pointers[first_sheet_title]
                prev_balance_cell = sheet.cell(row=pointer[1] - 1, column=pointer[0] + len(self.TRANSACTIONS_COLS) - 1)
                continue

            if start:
                pointer = sheet_pointers[sheet.title]
                amount_cell = sheet.cell(row=self.transaction_section[0][1] + 3,
                                         column=pointer[0] + len(self.TRANSACTIONS_COLS) - 3)

                new_balance_cell = sheet.cell(row=self.transaction_section[0][1] + 3,
                                              column=pointer[0] + len(self.TRANSACTIONS_COLS) - 1)
                new_balance_cell.value = '=SUM(\'{0}\'!{1},{2})'.format(prev_balance_cell.parent.title,
                                                                        prev_balance_cell.coordinate,
                                                                        amount_cell.coordinate)

                new_balance_cell.number_format = '#,###.00 [$AR$];[RED]-#,###.00 [$AR$]'
                new_balance_cell.font = self.grey_font
                new_balance_cell.alignment = Alignment(horizontal="justify")

                pointer = sheet_pointers[sheet.title]

                if pointer[1] > (self.transaction_section[0][1] + 3):
                    adjust = 1
                else:
                    adjust = 0

                prev_balance_cell = sheet.cell(row=pointer[1] - adjust,
                                               column=pointer[0] + len(self.TRANSACTIONS_COLS) - 1)

    def _calc_applicable_months(self, transaction, num_months):
        applicable_months = []
        next_date = datetime.strptime(transaction[5], "%Y-%m-%d %H:%M:%S")

        num_reps = transaction[7]

        # If the number of repetitions is not defined explicitly, check whether the quotes (#/#) can define them
        try:
            if num_reps == "":
                quote_num, total_quotes = [int(x) for x in transaction[8].split('/')]
                num_reps = (total_quotes - quote_num) + 1
            else:
                num_reps = int(num_reps)
        except ValueError:
            # The quotes field is not defined. Initialize the number of reps to the maximum possible.
            num_reps, total_quotes = num_months, None

        # Save the step interval. If the interval is an integer, cast it to int. If it is not an integer and the field
        # is empty, set the interval to 1.
        try:
            interval = transaction[6]
            interval = step = int(interval)
        except ValueError as e:
            if transaction[6] == "":
                interval = step = 1
            else:
                step = 2

        # The first applicable month is always the month of the start date
        applicable_months.append("{0} {1}".format(calendar.month_name[next_date.month], next_date.year))
        num_reps -= 1
        try:
            num_months -= step
            pass
        except:
            pass

        while num_reps > 0 and num_months > 0:
            next_date, increment = self._incr_month(next_date, interval)
            applicable_months.append("{0} {1}".format(calendar.month_name[next_date.month], next_date.year))
            num_reps -= 1
            num_months -= increment

        return applicable_months

    @staticmethod
    def _incr_month(start_date, interval):
        increment = interval
        try:
            next_date = start_date + relativedelta(months=+interval)
        except Exception as e:
            if (interval == "Uneven months" and start_date.month % 2) or \
                    (interval == "Even months" and not start_date.month % 2):
                increment = 2
                next_date = start_date + relativedelta(months=+increment)
            else:
                increment = 1
                next_date = start_date + relativedelta(months=+increment)

        return next_date, increment

    def _load_db(self):
        self.db_connection = sqlite3.connect(':memory:')
        self.db_cursor = self.db_connection.cursor()
        self.db_cursor.execute(
            "CREATE TABLE transactions (bank text, description text, subsection text, currency text, "
            "amount real, date numeric, interval text, reps text, quotes text)")

        # Read the Excel file and fill the database
        start_row = self._find_row(self.BALANCES_COLS) + 1
        for row in self.current_sheet.iter_rows(min_row=start_row,
                                                max_col=len(self.BALANCES_COLS),
                                                max_row=self.num_rows):
            if row[0].value is not None:
                cell_values = ["'{0}'".format(cell.value) for cell in row]
                insert_string = ','.join(cell_values)
                insert_string = insert_string.replace("None", "")
                self.db_cursor.execute("INSERT INTO transactions VALUES ({0})".format(insert_string))
            else:
                break

        self.db_connection.commit()

    def _find_row(self, columns):
        for index, row in enumerate(self.current_sheet.iter_rows(min_row=1,
                                                                 max_col=len(columns),
                                                                 max_row=self.num_rows)):
            row_headers = [str(cell.value) for cell in row]
            if sorted(row_headers) == sorted(columns):
                return index + 1

    def _print_frame(self, account_names, account_cols, transaction_cols, years, months):
        output = Workbook()
        start_row = 1
        start_column = 1
        # account_names = sorted(self.accounts_current.keys())
        self.accounts_section.append([start_column, start_row])
        self.accounts_section.append([len(self.ACCOUNTS_COLS), start_row + len(account_names)])

        # Create sheets for each month until the end of the prognosis
        now = date.today()
        then = now + relativedelta(months=+months, years=+years)
        for monthly in rrule.rrule(rrule.MONTHLY, dtstart=now, until=then):
            month_sheet = output.create_sheet(title="{0} {1}".format(calendar.month_name[monthly.month], monthly.year))

            # TODO: DRY

            # # Print bank headers
            # for col_nr, header in enumerate(account_cols):
            #     next_column = col_nr + 1
            #     cell = month_sheet.cell(row=start_row, column=next_column)
            #     cell.font = self.header_font
            #     cell.value = header
            #
            # # Print bank names
            # for row_nr, account_name in enumerate(account_names):
            #     next_row = start_row + row_nr + 1
            #     cell = month_sheet.cell(row=next_row, column=start_column)
            #     cell.font = self.text_font
            #     cell.value = account_name

            # Reserve section for transactions for each bank
            next_row = start_row
            for index, account_name in enumerate(account_names):
                next_column = 1 + index * (len(self.TRANSACTIONS_COLS) + 1)
                cell = month_sheet.cell(row=next_row, column=next_column)
                cell.font = Font(name="Calibri", size=14, bold=True)
                cell.value = account_name
                cell.alignment = Alignment(horizontal="justify")

                # Print transaction headers
                next_row += 2
                for col_nr, header in enumerate(transaction_cols):
                    cell = month_sheet.cell(row=next_row, column=next_column + col_nr)
                    cell.font = self.header_font
                    cell.value = header

                # Print dummy values for empty transactions later on.
                next_row += 1
                for col_nr, header in enumerate(transaction_cols[:-1]):
                    cell = month_sheet.cell(row=next_row, column=next_column + col_nr)
                    cell.font = Font(name="Calibri", size=11, color="808080")
                    cell.value = "N/A"
                    cell.alignment = Alignment(horizontal="justify")

                next_row -= 3

        self.transaction_section.append([start_column, next_row])

        # Remove the empty default sheet at the beginning
        output.remove_sheet(output.active)

        return output

    def __del__(self):
        self.db_connection.close()


class PrognosisWorkbook(BudgetWorkbook):
    def __init__(self):
        super(PrognosisWorkbook, self).__init__()


def main():
    calculator = BudgetCalc()
    calculator.read_input(sys.argv[1])
    calculator.save_prognosis(sys.argv[2], years=int(sys.argv[3]), months=int(sys.argv[4]))


if __name__ == '__main__':
    main()
